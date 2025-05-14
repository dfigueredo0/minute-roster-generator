import os

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment

from constants import *
from utils import * 

def create_table(writer, df, start_row, start_column, title, sheet_name="Sheet1"):
    """
    Inserts a DataFrame into an Excel worksheet at a specified location.
    Optionally adds a title row and a formatted header row for roll tracking.
    """
    # Write the DataFrame into the sheet starting from the specified row/column, skipping index and header
    df.to_excel(writer, sheet_name=sheet_name, startrow=start_row + 2, startcol=start_column, index=False, header=False)
    ws = writer.sheets[sheet_name]

    if title:
        # Format and center the title across the top of the table
        col_letter_start = get_column_letter(start_column + 1)
        col_letter_end = get_column_letter(start_column + len(df.columns))
        title_cell = ws.cell(row=start_row + 1, column=start_column + 1)
        title_cell.value = title
        title_cell.font = Font(bold=True)
        title_cell.alignment = Alignment(horizontal="center")
        ws.merge_cells(f"{col_letter_start}{start_row + 1}:{col_letter_end}{start_row + 1}")

    # Create and format the header row
    header_row = start_row + 2
    ws.cell(row=header_row, column=start_column + 1).value = "Officers"
    ws.merge_cells(start_row=header_row, start_column=start_column + 1, end_row=header_row, end_column=start_column + 2)
    ws.cell(row=header_row, column=start_column + 3).value = "Opening Roll"
    ws.cell(row=header_row, column=start_column + 4).value = "Closing Roll"

    for col_offset in range(4):
        cell = ws.cell(row=header_row, column=start_column + 1 + col_offset)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Center align values such as "P" or "E" (for Present/Excused)
    for row in ws.iter_rows(
        min_row=header_row + 1,
        max_row=header_row + 1 + len(df),
        min_col=start_column + 1,
        max_col=start_column + 4,
    ):
        for cell in row:
            if cell.value in ("P", "E"):
                cell.alignment = Alignment(horizontal="center")

    # Auto-fit the column widths for readability
    auto_adjust_column_widths(ws, df, start_row, start_column)


def create_segmented_table(writer, segments, start_row, start_col, title=None, sheet_name="Sheet1"):
    """
    Writes multiple dataframes into a single worksheet, one after another, 
    each optionally with a title and merged headers based on column names.
    """
    # Ensure the sheet exists
    ws = writer.sheets.get(sheet_name)
    if not ws:
        pd.DataFrame().to_excel(writer, sheet_name=sheet_name, index=False)
        ws = writer.sheets[sheet_name]

    current_row = start_row + 1

    for segment_title, df, headers in segments:
        if df.empty:
            continue  # Skip empty tables

        if segment_title:
            # Merge cells and insert the segment title
            col_start = get_column_letter(start_col + 1)
            col_end = get_column_letter(start_col + len(headers))
            ws.merge_cells(f"{col_start}{current_row}:{col_end}{current_row}")
            cell = ws.cell(row=current_row, column=start_col + 1)
            cell.value = segment_title
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            current_row += 1

        # Normalize headers for comparison
        normalized_headers = [h.strip().lower() for h in headers]

        # Handle common header patterns with merged cells
        if normalized_headers == ["officers", "full name", "opening roll", "closing roll"]:
            ws.merge_cells(start_row=current_row, start_column=start_col + 1, end_row=current_row, end_column=start_col + 2)
            ws.merge_cells(start_row=current_row, start_column=start_col + 3, end_row=current_row, end_column=start_col + 4)
            ws.cell(row=current_row, column=start_col + 1, value="Officers").font = Font(bold=True)
            ws.cell(row=current_row, column=start_col + 1).alignment = Alignment(horizontal="center")
            ws.cell(row=current_row, column=start_col + 3, value="Roll").font = Font(bold=True)
            ws.cell(row=current_row, column=start_col + 3).alignment = Alignment(horizontal="center")
            current_row += 1

        elif normalized_headers == ["officers", "full name", "roll"] or normalized_headers == ["others", "roll"]:
            ws.merge_cells(start_row=current_row, start_column=start_col + 1, end_row=current_row, end_column=start_col + 2)
            ws.cell(row=current_row, column=start_col + 1, value=headers[0]).font = Font(bold=True)
            ws.cell(row=current_row, column=start_col + 1).alignment = Alignment(horizontal="center")
            ws.cell(row=current_row, column=start_col + 3, value="Roll").font = Font(bold=True)
            ws.cell(row=current_row, column=start_col + 3).alignment = Alignment(horizontal="center")
            current_row += 1

        # Write the actual column headers 
        # TODO: Enable some check so Events, Finance, IOC, and Bylaws don't get double headers
        for i, label in enumerate(headers):
            cell = ws.cell(row=current_row, column=start_col + 1 + i)
            cell.value = label
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
        current_row += 1

        # Write the data rows, aligning "P"/"E" to center
        for row_values in df.values:
            for i, value in enumerate(row_values):
                cell = ws.cell(row=current_row, column=start_col + 1 + i)
                cell.value = value
                if value in ("P", "E"):
                    cell.alignment = Alignment(horizontal="center")
            current_row += 1

        auto_adjust_column_widths(ws, df, start_row, start_col)

    return current_row


def create_segment(*args, titles=None):
    """
    Creates a standardized list of (title, dataframe, headers) segments for table generation.
    """
    segments = []
    titles = titles or [""] * len(args)

    for i, df in enumerate(args):
        if not isinstance(df, pd.DataFrame):
            raise ValueError(f"Expected DataFrame, got {type(df)}")
        title = titles[i] if i < len(titles) else ""
        headers = list(df.columns)
        segments.append((title, df, headers))

    return segments


def create_new_members_df(count=6):
    """
    Returns a new DataFrame template for new members with default roll values.
    """
    return pd.DataFrame([{
        "Last Name": "",
        "First Name": "",
        "Opening Roll": "P",
        "Closing Roll": "P"
    } for _ in range(count)])


def create_others_df(count=4):
    """
    Returns a DataFrame for non-members with a combined 'Others' name field and roll status.
    """
    df = pd.DataFrame([{
        "Last Name": "",
        "First Name": "",
        "Roll": "P"
    } for _ in range(count)])
    
    return df.assign(Others=df["Last Name"] + " " + df["First Name"])[["Others", "Roll"]]


def process_advisors(df):
    """
    Processes an advisor DataFrame and returns a formatted version
    with roll assignments and staff roles sorted by defined rank.
    """
    return (
        df.assign(
            Chapter_Staff=df["First Name"] + " " + df["Last Name"],
            Opening_Roll=df["Current Office"].apply(
                lambda x: "E" if x in ["Chapter Advisor", "Asst. Chapter Advisor"] else "P"
            ),
            Closing_Roll=lambda d: d["Opening_Roll"],
            Role=df["Current Office"],
            Rank=df["Current Office"].apply(
                lambda role: next((i for i, r in enumerate(advisors) if r.lower() == role.lower()), len(advisors))
            )
        )
        .sort_values("Rank")
        .loc[:, ["Chapter_Staff", "Opening_Roll", "Closing_Roll", "Role"]]
        .rename(columns={
            "Chapter_Staff": "Chapter Staff",
            "Opening_Roll": "Opening Roll",
            "Closing_Roll": "Closing Roll"
        })
    )


def create_brothers_df(active_df):
    """
    Filters out officer rows and prepares a table for non-officer brothers with roll statuses.
    """
    df = active_df[~active_df['Current Office'].str.contains('|'.join(officers), na=False)].copy()
    return df.assign(
        Brothers=df["Last Name"],
        **{"Opening Roll": "P", "Closing Roll": "P"}
    )[["Brothers", "First Name", "Opening Roll", "Closing Roll"]]


def create_roster(writer, xlsx_output_dir, active_df, advisor_df):
    """
    Orchestrates the creation and formatting of an Excel workbook roster.
    It includes executive officers, advisors, members, and committee tables.
    """
    pd.DataFrame().to_excel(writer, sheet_name='Sheet1', index=False)
    output_path = os.path.join(xlsx_output_dir, 'Officer Roster and Minutes Rosters.xlsx')

    # Generate committee DataFrames
    executive_df = create_df(active_df, exec)
    events_df = create_df(active_df, events)
    finance_df = create_df(active_df, ['Asst. Tau', 'Sigma'])
    ioc_df = create_df(active_df, ['Beta', 'Theta One', 'Theta Two', 'Theta Three', 'Sigma'])
    bylaws_df = create_df(active_df, ['Sigma', 'Sigma'])
    officers_df = create_df(active_df, officers)

    # Generate additional tables
    brothers_df = create_brothers_df(active_df)
    new_members_df = create_new_members_df()
    others_df = create_others_df()
    advisor_df = process_advisors(advisor_df)

    # Grouped table segments
    segments = [
        ("Officers", officers_df, ["Officers", "Full Name", "Opening Roll", "Closing Roll"]),
        (" ", advisor_df, ["Chapter Staff", "Opening Roll", "Closing Roll", "Role"]),
        ("NEW MEMBERS", new_members_df, ["Last Name", "First Name", "Opening Roll", "Closing Roll"])
    ]
    chapter_segments = [
        ("", officers_df, ["Officers", "Full Name", "Opening Roll", "Closing Roll"]),
        ("", brothers_df, ["Brothers", "First Name", "Opening Roll", "Closing Roll"]),
        ("", advisor_df, ["Chapter Staff", "Opening Roll", "Closing Roll", "Role"]),
    ]

    # Main tables
    create_table(writer, executive_df, 0, table_positions['EXECUTIVE COUNCIL COMMITTEE'], 'EXECUTIVE COUNCIL COMMITTEE')
    create_segmented_table(writer, chapter_segments, len(executive_df) + 3, table_positions['EXECUTIVE COUNCIL COMMITTEE'])
    create_segmented_table(writer, segments, 0, table_positions['HOUSE'])

    # Write all other committee segments dynamically
    committees = [
        ("EVENTS COMMITTEE", events_df),
        ("FINANCE COMMITTEE", finance_df),
        ("INTERNAL OPERATIONS COMMITTEE", ioc_df),
        ("BYLAWS COMMITTEE", bylaws_df)
    ]

    row_offset = 0
    for name, df in committees:
        segment = create_segment(df, others_df, titles=[name])
        row_offset = create_segmented_table(writer, segment, row_offset, table_positions[name])

    # Save final workbook
    writer.book.save(output_path)